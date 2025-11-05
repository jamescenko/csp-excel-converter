from flask import Flask, request, send_file, jsonify
from openpyxl import load_workbook
from openpyxl.styles import numbers
from openpyxl.cell.cell import MergedCell
from datetime import datetime
import io
import os
import traceback

app = Flask(__name__)

@app.route('/health', methods=['GET'])
def health():
    return jsonify({
        "status": "healthy",
        "message": "CSP Excel Service Running",
        "timestamp": datetime.now().isoformat()
    })

@app.route('/populate-excel', methods=['POST'])
def populate_excel():
    try:
        data = request.json
        if not data:
            return jsonify({"error": "No data received"}), 400
        
        print(f"\n{'='*80}")
        print(f"📥 RAW DATA RECEIVED:")
        print(f"Top-level keys: {list(data.keys())}")
        
        # Extract data
        summary_data = data.get('summary', {})
        exchange_rate = float(data.get('exchangeRate', 1.39))
        period_from = data.get('reportPeriodFrom', '01-01-2025')
        period_to = data.get('reportPeriodTo', '31-03-2025')
        
        # CRITICAL: Get regions from BOTH possible paths
        regions_data = data.get('regions', [])
        if not regions_data:
            regions_data = summary_data.get('regions', [])
        
        print(f"\n🔍 DATA ANALYSIS:")
        print(f"   Summary keys: {list(summary_data.keys())}")
        print(f"   Regions found: {len(regions_data)}")
        if regions_data:
            print(f"   First region: {regions_data[0].get('code', 'NO CODE')} - {len(regions_data[0].get('childDetails', []))} children")
        print(f"   Exchange rate: {exchange_rate}")
        print(f"{'='*80}\n")
        
        # Load template
        template_path = 'CSP_Automate_Template.xlsx'
        if not os.path.exists(template_path):
            return jsonify({"error": "Template not found"}), 500
        
        wb = load_workbook(template_path)
        print(f"✓ Template loaded: {len(wb.sheetnames)} sheets")
        
        results = {
            "summary_updated": False,
            "regions_processed": 0,
            "regions_skipped": [],
            "children_written": 0,
            "errors": []
        }
        
        # Currency format
        currency_format = '#,##0.00'
        
        # ============================================
        # POPULATE SUMMARY SHEET
        # ============================================
        try:
            ws = wb['Summary']
            print(f"\n📊 POPULATING SUMMARY...")
            
            # Period & Exchange Rate
            ws['B4'] = period_from
            ws['B5'] = period_to
            ws['H4'] = exchange_rate
            ws['H4'].number_format = '#,##0.00'
            
            # Regular Support Section (Rows 20-25)
            ws['C20'] = int(summary_data.get('totalChildren', 0))
            ws['D20'] = float(summary_data.get('foodDistCAD', 0))
            ws['D20'].number_format = currency_format
            ws['E20'] = float(summary_data.get('foodDistUSD', 0))
            ws['E20'].number_format = currency_format
            
            ws['D22'] = float(summary_data.get('salaryCAD', 0))
            ws['D22'].number_format = currency_format
            ws['E22'] = float(summary_data.get('salaryUSD', 0))
            ws['E22'].number_format = currency_format
            
            ws['D24'] = float(summary_data.get('incentiveCAD', 0))
            ws['D24'].number_format = currency_format
            ws['E24'] = float(summary_data.get('incentiveUSD', 0))
            ws['E24'].number_format = currency_format
            
            # D25 and E25 SUBTOTALS
            d25 = float(summary_data.get('foodDistCAD', 0)) + float(summary_data.get('salaryCAD', 0)) + float(summary_data.get('incentiveCAD', 0))
            e25 = float(summary_data.get('foodDistUSD', 0)) + float(summary_data.get('salaryUSD', 0)) + float(summary_data.get('incentiveUSD', 0))
            ws['D25'] = d25
            ws['D25'].number_format = currency_format
            ws['E25'] = e25
            ws['E25'].number_format = currency_format
            
            # Additional Support (Rows 28-30)
            ws['D28'] = float(summary_data.get('familyCAD', 0))
            ws['D28'].number_format = currency_format
            ws['E28'] = float(summary_data.get('familyUSD', 0))
            ws['E28'].number_format = currency_format
            
            ws['D29'] = float(summary_data.get('medicalCAD', 0))
            ws['D29'].number_format = currency_format
            ws['E29'] = float(summary_data.get('medicalUSD', 0))
            ws['E29'].number_format = currency_format
            
            # D30 and E30 SUBTOTALS
            d30 = float(summary_data.get('familyCAD', 0)) + float(summary_data.get('medicalCAD', 0))
            e30 = float(summary_data.get('familyUSD', 0)) + float(summary_data.get('medicalUSD', 0))
            ws['D30'] = d30
            ws['D30'].number_format = currency_format
            ws['E30'] = e30
            ws['E30'].number_format = currency_format
            
            # D32 and E32 GRAND TOTALS
            ws['D32'] = d25 + d30
            ws['D32'].number_format = currency_format
            ws['E32'] = e25 + e30
            ws['E32'].number_format = currency_format
            
            # Cross Check Section (Rows 36-43)
            ws['C36'] = int(summary_data.get('totalChildren', 0))
            ws['C37'] = int(summary_data.get('newChildrenCount', 0))
            
            ws['C40'] = float(summary_data.get('foodDistCAD', 0))
            ws['C40'].number_format = currency_format
            ws['D40'] = float(summary_data.get('foodDistUSD', 0))
            ws['D40'].number_format = currency_format
            
            ws['C41'] = float(summary_data.get('salaryCAD', 0))
            ws['C41'].number_format = currency_format
            ws['D41'] = float(summary_data.get('salaryUSD', 0))
            ws['D41'].number_format = currency_format
            
            ws['C42'] = float(summary_data.get('incentiveCAD', 0))
            ws['C42'].number_format = currency_format
            ws['D42'] = float(summary_data.get('incentiveUSD', 0))
            ws['D42'].number_format = currency_format
            
            # C43 and D43 Subtotals
            c43 = float(summary_data.get('foodDistCAD', 0)) + float(summary_data.get('salaryCAD', 0)) + float(summary_data.get('incentiveCAD', 0))
            d43 = float(summary_data.get('foodDistUSD', 0)) + float(summary_data.get('salaryUSD', 0)) + float(summary_data.get('incentiveUSD', 0))
            ws['C43'] = c43
            ws['C43'].number_format = currency_format
            ws['D43'] = d43
            ws['D43'].number_format = currency_format
            
            # Admin Fee & Gifts (Rows 47-51)
            ws['C47'] = float(summary_data.get('familyCAD', 0))
            ws['C47'].number_format = currency_format
            ws['D47'] = float(summary_data.get('familyUSD', 0))
            ws['D47'].number_format = currency_format
            
            ws['C48'] = float(summary_data.get('medicalCAD', 0))
            ws['C48'].number_format = currency_format
            ws['D48'] = float(summary_data.get('medicalUSD', 0))
            ws['D48'].number_format = currency_format
            
            ws['C49'] = d30
            ws['C49'].number_format = currency_format
            ws['D49'] = e30
            ws['D49'].number_format = currency_format
            
            ws['C51'] = c43 + d30
            ws['C51'].number_format = currency_format
            ws['D51'] = d43 + e30
            ws['D51'].number_format = currency_format
            
            # Summary Statistics (Rows 55-58)
            ws['B55'] = int(summary_data.get('totalChildren', 0))
            ws['G55'] = exchange_rate
            ws['G55'].number_format = '#,##0.00'
            
            ws['B56'] = int(summary_data.get('newChildrenCount', 0))
            total_usd = float(summary_data.get('totalUSD', 0))
            total_children = int(summary_data.get('totalChildren', 1))
            ws['G56'] = round(total_usd / total_children, 2) if total_children > 0 else 0
            ws['G56'].number_format = currency_format
            
            ws['B57'] = int(summary_data.get('totalChildren', 0))
            ws['G57'] = float(summary_data.get('totalCAD', 0))
            ws['G57'].number_format = currency_format
            
            ws['B58'] = len(regions_data)
            ws['G58'] = float(summary_data.get('totalUSD', 0))
            ws['G58'].number_format = currency_format
            
            results['summary_updated'] = True
            print(f"   ✅ Summary done")
            
        except Exception as e:
            error_msg = f"Summary error: {str(e)}"
            results['errors'].append(error_msg)
            print(f"   ❌ {error_msg}")
            print(traceback.format_exc())
        
        # ============================================
        # POPULATE REGION SHEETS
        # ============================================
        print(f"\n📍 PROCESSING {len(regions_data)} REGIONS...")
        
        for idx, region in enumerate(regions_data):
            region_code = str(region.get('code', '')).strip().upper()
            
            print(f"\n[{idx+1}/{len(regions_data)}] {region_code}")
            
            if not region_code:
                results['regions_skipped'].append('UNKNOWN')
                continue
            
            # Find sheet
            sheet_found = None
            for sheet_name in wb.sheetnames:
                if sheet_name.upper() == region_code:
                    sheet_found = sheet_name
                    break
            
            if not sheet_found:
                results['regions_skipped'].append(f"{region_code}")
                print(f"   ⚠️ Skip: Not in template")
                continue
            
            try:
                ws = wb[sheet_found]
                
                # Period & Exchange Rate
                ws['B4'] = period_from
                ws['B5'] = period_to
                ws['H4'] = exchange_rate
                ws['H4'].number_format = '#,##0.00'
                
                # Wire & Location
                ws['B10'] = str(region.get('wireId', ''))
                ws['B11'] = region.get('caseworker', 'Unknown')
                ws['B13'] = region.get('beneficiary', '')
                ws['B14'] = region.get('region', region_code)
                ws['B15'] = region.get('city', '')
                
                # Regular Support (Rows 20-24)
                ws['C20'] = int(region.get('children', 0))
                ws['D20'] = float(region.get('foodDistCAD', 0))
                ws['D20'].number_format = currency_format
                ws['E20'] = float(region.get('foodDistUSD', 0))
                ws['E20'].number_format = currency_format
                
                ws['D22'] = float(region.get('salaryCAD', 0))
                ws['D22'].number_format = currency_format
                ws['E22'] = float(region.get('salaryUSD', 0))
                ws['E22'].number_format = currency_format
                
                ws['D24'] = float(region.get('incentiveCAD', 0))
                ws['D24'].number_format = currency_format
                ws['E24'] = float(region.get('incentiveUSD', 0))
                ws['E24'].number_format = currency_format
                
                # D25/E25 SUBTOTALS
                d25 = float(region.get('foodDistCAD', 0)) + float(region.get('salaryCAD', 0)) + float(region.get('incentiveCAD', 0))
                e25 = float(region.get('foodDistUSD', 0)) + float(region.get('salaryUSD', 0)) + float(region.get('incentiveUSD', 0))
                ws['D25'] = d25
                ws['D25'].number_format = currency_format
                ws['E25'] = e25
                ws['E25'].number_format = currency_format
                
                # Additional Support (Rows 28-30)
                ws['D28'] = float(region.get('familyCAD', 0))
                ws['D28'].number_format = currency_format
                ws['E28'] = float(region.get('familyUSD', 0))
                ws['E28'].number_format = currency_format
                
                ws['D29'] = float(region.get('medicalCAD', 0))
                ws['D29'].number_format = currency_format
                ws['E29'] = float(region.get('medicalUSD', 0))
                ws['E29'].number_format = currency_format
                
                # D30/E30 SUBTOTALS
                d30 = float(region.get('familyCAD', 0)) + float(region.get('medicalCAD', 0))
                e30 = float(region.get('familyUSD', 0)) + float(region.get('medicalUSD', 0))
                ws['D30'] = d30
                ws['D30'].number_format = currency_format
                ws['E30'] = e30
                ws['E30'].number_format = currency_format
                
                # D32/E32 GRAND TOTALS
                ws['D32'] = d25 + d30
                ws['D32'].number_format = currency_format
                ws['E32'] = e25 + e30
                ws['E32'].number_format = currency_format
                
                print(f"   💰 Totals: CAD ${d25 + d30:,.2f}, USD ${e25 + e30:,.2f}")
                
                # ============================================
                # Children Table (Row 36+) - FIXED FOR MERGED CELLS
                # ============================================
                child_details = region.get('childDetails', [])
                print(f"   👶 Children: {len(child_details)}")
                
                # Clear old data (SKIP MERGED CELLS)
                for row in range(36, 201):
                    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
                        cell = ws[f'{col}{row}']
                        if not isinstance(cell, MergedCell):
                            ws[f'{col}{row}'] = None
                
                # Write children
                if child_details and len(child_details) > 0:
                    for child_idx, child in enumerate(child_details):
                        row_num = 36 + child_idx
                        
                        csp_id = str(child.get('cspId', '')).strip()
                        child_name = str(child.get('childName', child.get('name', ''))).strip()
                        
                        # Financial data
                        food_usd = float(child.get('foodDistUSD', 0) if child.get('foodDistUSD', 0) else child.get('foodAmount', 0))
                        medical = float(child.get('medicalGifts', 0))
                        family_gift = float(child.get('familyGifts', 0))
                        total = round(food_usd + medical + family_gift, 2)
                        
                        # Write to cells
                        ws[f'A{row_num}'] = csp_id
                        ws[f'B{row_num}'] = child_name
                        ws[f'C{row_num}'] = food_usd
                        ws[f'C{row_num}'].number_format = currency_format
                        ws[f'D{row_num}'] = medical
                        ws[f'D{row_num}'].number_format = currency_format
                        ws[f'E{row_num}'] = family_gift
                        ws[f'E{row_num}'].number_format = currency_format
                        ws[f'F{row_num}'] = total
                        ws[f'F{row_num}'].number_format = currency_format
                        ws[f'G{row_num}'] = ''
                        
                        results['children_written'] += 1
                        
                        if child_idx == 0:
                            print(f"      ✍️ First: {csp_id} | {child_name} | ${food_usd:.2f}")
                    
                    print(f"      ✅ Wrote {len(child_details)} children")
                else:
                    print(f"      ⚠️ No children data")
                
                results['regions_processed'] += 1
                print(f"   ✅ Region done")
                
            except Exception as e:
                error_msg = f"{region_code}: {str(e)}"
                results['errors'].append(error_msg)
                results['regions_skipped'].append(region_code)
                print(f"   ❌ {str(e)}")
                print(traceback.format_exc())
        
        # SAVE
	if 'Summary' in wb.sheetnames:
    		wb.active = wb['Summary']
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        print(f"\n{'='*80}")
        print(f"✅ FINAL RESULTS:")
        print(f"   Summary: {'✓' if results['summary_updated'] else '✗'}")
        print(f"   Regions: {results['regions_processed']}/{len(regions_data)}")
        print(f"   Children: {results['children_written']}")
        if results['regions_skipped']:
            print(f"   Skipped: {', '.join(results['regions_skipped'][:10])}")
        print(f"{'='*80}\n")
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'CSP_Wiring_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        )
    
    except Exception as e:
        print(f"❌ CRITICAL: {str(e)}")
        print(traceback.format_exc())
        return jsonify({"error": str(e)}), 500

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 8080))
    app.run(host='0.0.0.0', port=port)
